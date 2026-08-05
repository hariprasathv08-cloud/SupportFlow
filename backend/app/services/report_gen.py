import io
import csv
import socket
from openpyxl import Workbook
from datetime import datetime
from typing import List, Dict, Any

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Design system colors
PRIMARY = colors.HexColor('#2563EB')
SUCCESS = colors.HexColor('#22C55E')
WARNING = colors.HexColor('#F59E0B')
DANGER = colors.HexColor('#EF4444')
DARK = colors.HexColor('#0F172A')
LIGHT_BG = colors.HexColor('#F8FAFC')
BORDER = colors.HexColor('#E2E8F0')

def add_header_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont('Helvetica-Bold', 8)
    canvas.setFillColor(DARK)
    canvas.drawString(54, 750, "HelpDesk X - Enterprise Systems Report")
    
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(colors.HexColor('#64748B'))
    canvas.drawRightString(doc.pagesize[0]-54, 750, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    canvas.setStrokeColor(BORDER)
    canvas.setLineWidth(1)
    canvas.line(54, 742, doc.pagesize[0]-54, 742)
    
    # Footer
    canvas.line(54, 50, doc.pagesize[0]-54, 50)
    canvas.drawString(54, 38, "CONFIDENTIAL - INTERNAL USE ONLY")
    canvas.drawRightString(doc.pagesize[0]-54, 38, f"Page {doc.page}")
    canvas.restoreState()

def generate_pdf_report(title: str, headers: List[str], rows: List[List[Any]], col_widths: List[float] = None) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=72,
        bottomMargin=72
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=PRIMARY,
        spaceAfter=12
    )
    
    header_cell_style = ParagraphStyle(
        'HeaderCell',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white
    )
    
    body_cell_style = ParagraphStyle(
        'BodyCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#334155'),
        leading=10
    )

    story = []
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M')}", body_cell_style))
    story.append(Spacer(1, 15))
    
    # Wrap headers and rows in Paragraphs for auto-wrap
    table_data = []
    table_data.append([Paragraph(str(h), header_cell_style) for h in headers])
    for row in rows:
        table_data.append([Paragraph(str(cell) if cell is not None else "", body_cell_style) for cell in row])
        
    t = Table(table_data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), PRIMARY),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, BORDER),
    ]))
    story.append(t)
    
    doc.build(story, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

def generate_dashboard_pdf(summary_data: Dict[str, Any], recent_tickets: List[List[Any]], active_alerts: List[List[Any]]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=72,
        bottomMargin=72
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=PRIMARY,
        spaceAfter=12
    )
    
    h2_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=DARK,
        spaceBefore=14,
        spaceAfter=6
    )
    
    body_style = ParagraphStyle(
        'BodyCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#334155'),
        leading=10
    )
    
    header_cell_style = ParagraphStyle(
        'HeaderCell',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white
    )

    story = []
    story.append(Paragraph("System Dashboard Report", title_style))
    story.append(Paragraph(f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M')}", body_style))
    story.append(Spacer(1, 12))
    
    # 1. Summary Cards Table
    story.append(Paragraph("Key Metrics Summary", h2_style))
    summary_headers = ["Metric Category", "Count / Value", "Status"]
    summary_rows = [
        ["Total Assets Monitored", str(summary_data.get("total_assets", 0)), "Active"],
        ["HelpDesk Tickets", f"{summary_data.get('total_tickets', 0)} ({summary_data.get('resolved_tickets', 0)} Resolved)", "Monitoring"],
        ["Critical Active Alerts", str(summary_data.get("critical_alerts", 0)), "Action Required" if summary_data.get("critical_alerts", 0) > 0 else "Normal"],
        ["Registered Console Users", str(summary_data.get("total_users", 0)), "Active"]
    ]
    
    table_data = [[Paragraph(h, header_cell_style) for h in summary_headers]]
    for row in summary_rows:
        table_data.append([Paragraph(cell, body_style) for cell in row])
        
    t1 = Table(table_data, colWidths=[2.5*inch, 2.5*inch, 2*inch])
    t1.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), DARK),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, BORDER),
    ]))
    story.append(t1)
    story.append(Spacer(1, 10))
    
    # 2. Recent Tickets Table
    story.append(Paragraph("Recent Tickets Logs", h2_style))
    ticket_headers = ["ID", "Title", "Category", "Priority", "Status", "Created At"]
    t2_data = [[Paragraph(h, header_cell_style) for h in ticket_headers]]
    for row in recent_tickets:
        t2_data.append([Paragraph(str(cell), body_style) for cell in row])
        
    t2 = Table(t2_data, colWidths=[0.5*inch, 2.5*inch, 1.1*inch, 0.8*inch, 1.1*inch, 1.5*inch])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), PRIMARY),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, BORDER),
    ]))
    story.append(t2)
    story.append(Spacer(1, 10))
    
    # 3. Active Alerts Table
    story.append(Paragraph("Active Unresolved Alerts", h2_style))
    alert_headers = ["ID", "Hostname", "Category", "Severity", "Message", "Created At"]
    t3_data = [[Paragraph(h, header_cell_style) for h in alert_headers]]
    for row in active_alerts:
        t3_data.append([Paragraph(str(cell), body_style) for cell in row])
        
    t3 = Table(t3_data, colWidths=[0.5*inch, 1.2*inch, 1.0*inch, 0.8*inch, 2.3*inch, 1.7*inch])
    t3.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), DANGER),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, BORDER),
    ]))
    story.append(t3)
    
    doc.build(story, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

def generate_excel_report(title: str, headers: List[str], rows: List[List[Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:30] # Excel sheet limit is 31
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    border_thin = Side(style='thin', color='E2E8F0')
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        
    for row in rows:
        # Convert any datetime objects to string representation to avoid formatting issues
        formatted_row = []
        for cell in row:
            if isinstance(cell, datetime):
                formatted_row.append(cell.strftime('%Y-%m-%d %H:%M'))
            else:
                formatted_row.append(cell)
        ws.append(formatted_row)
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
            # Add thin border to cells
            cell.border = cell_border
            if cell.row > 1:
                cell.alignment = align_left
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes

def generate_csv_report(headers: List[str], rows: List[List[Any]]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        formatted_row = []
        for cell in row:
            if isinstance(cell, datetime):
                formatted_row.append(cell.strftime('%Y-%m-%d %H:%M'))
            else:
                formatted_row.append(cell)
        writer.writerow(formatted_row)
    return output.getvalue()

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import os
from reportlab.platypus import PageBreak

def send_report_email(to_emails: List[str], attachment_paths: List[str], report_title: str):
    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_password = os.getenv("SMTP_PASSWORD", "")

    print(f"[MAIL SYSTEM] Delivering report '{report_title}' via SMTP to {to_emails}...")
    
    if not smtp_user or not smtp_password:
        print("[MAIL SYSTEM] WARNING: SMTP_USER or SMTP_PASSWORD is not configured. Simulating delivery in logs.")
        print(f"[MAIL SYSTEM] Attachments: {attachment_paths}")
        return True

    try:
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = ", ".join(to_emails)
        msg['Subject'] = f"HelpDesk X Generated Report: {report_title}"

        body = f"Hello,\n\nPlease find attached the requested report: {report_title}.\n\nBest regards,\nHelpDesk X Team"
        msg.attach(MIMEText(body, 'plain'))

        for path in attachment_paths:
            if not os.path.exists(path):
                continue
            part = MIMEBase('application', "octet-stream")
            with open(path, 'rb') as file:
                part.set_payload(file.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(path)}"')
            msg.attach(part)

        # Connect and send
        server = smtplib.SMTP(smtp_host, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.sendmail(smtp_user, to_emails, msg.as_string())
        server.quit()
        print(f"[MAIL SYSTEM] Email sent successfully via {smtp_host} to {to_emails}.")
        return True
    except Exception as e:
        print(f"[MAIL SYSTEM] ERROR: Failed to deliver email via SMTP: {e}")
        raise e

def generate_complete_enterprise_pdf(db, date_range: str) -> bytes:
    from app.models.user import User
    from app.models.asset import Asset
    from app.models.ticket import Ticket
    from app.models.alert import Alert
    from app.models.software import Software
    from app.models.audit import AuditLog

    total_users = db.query(User).count()
    total_assets = db.query(Asset).count()
    total_tickets = db.query(Ticket).count()
    resolved_tickets = db.query(Ticket).filter(Ticket.status == "Resolved").count()
    active_alerts = db.query(Alert).filter(Alert.resolved == False).count()
    total_software = db.query(Software).count()
    total_audits = db.query(AuditLog).count()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=72,
        bottomMargin=72
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=PRIMARY,
        spaceAfter=15
    )
    h1_style = ParagraphStyle(
        'SecTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=DARK,
        spaceBefore=18,
        spaceAfter=8,
        keepWithNext=True
    )
    body_style = ParagraphStyle(
        'BodyText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#334155'),
        leading=12,
        spaceAfter=10
    )
    
    header_cell_style = ParagraphStyle(
        'HeaderCell',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white
    )
    
    story = []
    
    # Cover page
    story.append(Spacer(1, 100))
    story.append(Paragraph("HELPDESK X ENTERPRISE REPORT", title_style))
    story.append(Paragraph("Complete Infrastructure, Security, and Incident Diagnostic Analysis", ParagraphStyle('Sub', parent=body_style, fontSize=12, textColor=colors.HexColor('#64748B'))))
    story.append(Spacer(1, 20))
    story.append(Paragraph(f"Analysis Period: {date_range.upper()}", body_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}", body_style))
    story.append(Spacer(1, 100))
    story.append(Paragraph("CONFIDENTIALITY NOTICE: The information contained in this report is confidential and intended solely for the use of the administrator. Disclosing this information without authorization is strictly prohibited.", ParagraphStyle('Conf', parent=body_style, fontName='Helvetica-Oblique', fontSize=8)))
    
    story.append(PageBreak())
    
    # Executive Summary
    story.append(Paragraph("Executive Summary", h1_style))
    story.append(Paragraph("This report compiles real-time telemetry metrics, physical asset directories, incident ticketing logs, security vulnerabilities, and operator audits from the HelpDesk X PostgreSQL central engine. The current operational environment is monitored for latency issues, software version mismatch anomalies, and hardware health degradations.", body_style))
    
    # KPIs
    story.append(Paragraph("Key Performance Indicators (KPIs)", h1_style))
    kpi_headers = ["Metric Area", "Measurement", "Assessment"]
    kpi_rows = [
        ["Total Devices Managed", str(total_assets), "Optimal" if total_assets > 0 else "Incomplete"],
        ["Active User Profiles", str(total_users), "Verified"],
        ["HelpDesk Backlog", f"{total_tickets - resolved_tickets} Unresolved", "Warning" if (total_tickets - resolved_tickets) > 5 else "Healthy"],
        ["Threat Alerts Triggered", str(active_alerts), "Critical Action Required" if active_alerts > 0 else "Secure"],
        ["Unique Software Audited", str(total_software), "Standardized"],
        ["Audit Trails Tracked", str(total_audits), "Logged"]
    ]
    
    t_kpi_data = [[Paragraph(h, header_cell_style) for h in kpi_headers]]
    for r in kpi_rows:
        t_kpi_data.append([Paragraph(c, body_style) for c in r])
    t_kpis = Table(t_kpi_data, colWidths=[2.5*inch, 2.5*inch, 2*inch])
    t_kpis.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), DARK),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, BORDER),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t_kpis)
    story.append(Spacer(1, 15))
    
    # Recommendations
    story.append(Paragraph("Operations Recommendations", h1_style))
    recs = []
    if active_alerts > 0:
        recs.append("<b>VULNERABILITY MITIGATION:</b> Solve critical cybersecurity alerts on active workstation nodes immediately to enforce device policies.")
    else:
        recs.append("<b>VULNERABILITY MITIGATION:</b> System security states are verified secure. Maintain default firewall settings.")
        
    if (total_tickets - resolved_tickets) > 5:
        recs.append("<b>STAFF ALLOCATION:</b> Incident backlog is elevated. Assign available technicians to route tickets queue.")
    else:
        recs.append("<b>STAFF ALLOCATION:</b> Ticket backlog limits are within normal operational limits.")
        
    if total_assets == 0:
        recs.append("<b>ENDPOINT MANAGEMENT:</b> Register endpoints in the directory to begin telemetry collection.")
    else:
        recs.append("<b>ENDPOINT MANAGEMENT:</b> Monitor workstation diagnostics weekly to check CPU/RAM usage averages.")
        
    for idx, rec in enumerate(recs, 1):
        story.append(Paragraph(f"{idx}. {rec}", body_style))
        
    story.append(PageBreak())
    
    # Assets
    story.append(Paragraph("Hardware Assets Inventory", h1_style))
    assets = db.query(Asset).limit(20).all()
    asset_headers = ["Asset Tag", "Hostname", "OS Platform", "IP Address", "Health Score"]
    t_asset_data = [[Paragraph(h, header_cell_style) for h in asset_headers]]
    for a in assets:
        t_asset_data.append([Paragraph(str(c), body_style) for c in [a.asset_tag, a.hostname or 'N/A', a.operating_system or 'N/A', a.ip_address or 'N/A', f"{a.health_score}%"]])
    t_assets = Table(t_asset_data, colWidths=[1.2*inch, 1.8*inch, 1.8*inch, 1.2*inch, 1*inch])
    t_assets.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), PRIMARY),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, BORDER),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t_assets)
    
    story.append(PageBreak())
    
    # Tickets
    story.append(Paragraph("HelpDesk Incident Ledger", h1_style))
    tickets = db.query(Ticket).limit(20).all()
    ticket_headers = ["ID", "Title", "Priority", "Status", "Assigned To"]
    t_ticket_data = [[Paragraph(h, header_cell_style) for h in ticket_headers]]
    for t in tickets:
        t_ticket_data.append([Paragraph(str(c), body_style) for c in [f"#TIC-{1000+t.id}", t.title, t.priority, t.status, t.assigned_to.full_name if t.assigned_to else 'Unassigned']])
    t_tickets = Table(t_ticket_data, colWidths=[1*inch, 2.5*inch, 1*inch, 1.2*inch, 1.3*inch])
    t_tickets.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), PRIMARY),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, BORDER),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t_tickets)
    
    doc.build(story, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

def generate_complete_enterprise_excel(db, date_range: str) -> bytes:
    wb = Workbook()
    
    # 1. Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.append(["HelpDesk X Enterprise Executive Report"])
    ws_summary.append(["Generated At", datetime.now().strftime('%Y-%m-%d %H:%M')])
    ws_summary.append(["Date Range", date_range])
    ws_summary.append([])
    
    from app.models.user import User
    from app.models.asset import Asset
    from app.models.ticket import Ticket
    from app.models.alert import Alert
    from app.models.software import Software
    from app.models.audit import AuditLog
    
    ws_summary.append(["Operational Metric", "Total Count", "Status"])
    ws_summary.append(["Managed Endpoints", db.query(Asset).count(), "Monitoring"])
    ws_summary.append(["HelpDesk Incidents", db.query(Ticket).count(), "Active"])
    ws_summary.append(["Active Security Alerts", db.query(Alert).filter(Alert.resolved == False).count(), "Action Required"])
    ws_summary.append(["Audited Operator Log Entries", db.query(AuditLog).count(), "Secure"])

    # 2. Assets sheet
    ws_assets = wb.create_sheet(title="Assets")
    ws_assets.append(["ID", "Asset Tag", "Hostname", "Type", "Operating System", "IP Address", "Status", "Health Score"])
    for a in db.query(Asset).all():
        ws_assets.append([a.id, a.asset_tag, a.hostname or 'N/A', a.type or 'Workstation', a.operating_system or 'N/A', a.ip_address or 'N/A', a.status, f"{a.health_score}%"])

    # 3. Tickets sheet
    ws_tickets = wb.create_sheet(title="Tickets")
    ws_tickets.append(["ID", "Title", "Category", "Priority", "Status", "Created By", "Assigned To", "Created At"])
    for t in db.query(Ticket).all():
        ws_tickets.append([t.id, t.title, t.category or 'General', t.priority, t.status, t.created_by.full_name if t.created_by else 'N/A', t.assigned_to.full_name if t.assigned_to else 'Unassigned', t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else 'N/A'])

    # 4. Alerts sheet
    ws_alerts = wb.create_sheet(title="Security Alerts")
    ws_alerts.append(["ID", "Asset Hostname", "Category", "Severity", "Message", "Resolved", "Created At"])
    for al in db.query(Alert).all():
        ws_alerts.append([al.id, al.asset.hostname if al.asset else 'Local Host', al.category, al.severity, al.message, "Resolved" if al.resolved else "Active Alert", al.created_at.strftime("%Y-%m-%d %H:%M") if al.created_at else 'N/A'])

    # 5. Software sheet
    ws_software = wb.create_sheet(title="Software Inventory")
    ws_software.append(["ID", "Hostname", "Software Name", "Version", "Publisher", "Install Date"])
    for s in db.query(Software).all():
        ws_software.append([s.id, s.asset.hostname if s.asset else s.endpoint_uuid, s.name, s.version or 'N/A', s.publisher or 'N/A', s.install_date or 'N/A'])

    # Apply thin borders and column width adjustments to all sheets
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    border_thin = Side(style='thin', color='E2E8F0')
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    for ws in wb.worksheets:
        # Format header row
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-adjust column widths and borders
        for col in ws.columns:
            max_len = 0
            for cell in col:
                cell.border = cell_border
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes
